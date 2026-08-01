"""Regenerates Sample_Dashboard_Data.xlsx/.csv (at the repo root) against the current
(2025-01-01 -> today, auto-growing) dataset. Mirrors the structure/styling of the original
file: a README sheet plus Unified Metrics, Campaigns, Adsets, Platform Metrics, Salesforce
CPV, Adobe Visits.

Run from anywhere with: python backend/build_sample_data.py
(uses its own throwaway SQLite file in the OS temp dir -- does not touch the app's real DB)
"""
import os
import sys
import tempfile
from datetime import date

sys.path.insert(0, os.path.dirname(__file__))
os.environ.setdefault("DASHBOARD_DB_PATH", os.path.join(tempfile.gettempdir(), "sample_data_build.db"))
if os.path.exists(os.environ["DASHBOARD_DB_PATH"]):
    os.remove(os.environ["DASHBOARD_DB_PATH"])

import pandas as pd
from openpyxl.styles import Font, PatternFill

from app import db

conn = db.get_connection()

campaigns = pd.read_sql_query("SELECT * FROM campaigns", conn)
adsets = pd.read_sql_query("SELECT * FROM adsets", conn)
platform_metrics = pd.read_sql_query(
    "SELECT metric_date, adset_id, platform, impressions, clicks, spend, conversions "
    "FROM platform_metrics ORDER BY metric_date, adset_id", conn
)
salesforce_cpv = pd.read_sql_query(
    "SELECT metric_date, ecd_code, cpv FROM salesforce_cpv ORDER BY metric_date, ecd_code", conn
)
adobe_visits = pd.read_sql_query(
    "SELECT metric_date, ecd_code, visits FROM adobe_visits ORDER BY metric_date, ecd_code", conn
)
unified = pd.read_sql_query("SELECT * FROM unified_metrics ORDER BY metric_date, ecd_code", conn)
conn.close()

today_str = date.today().isoformat()
readme_rows = [
    ("Sheet", "What it is"),
    ("Unified Metrics", "One row per adset per day — this is the joined table the chat assistant actually queries. Use this sheet to verify any number shown on the dashboard."),
    ("Campaigns", "One row per campaign: name and which ad platform it runs on."),
    ("Adsets", "One row per adset. ecd_code is the unique key used to join Salesforce and Adobe data below."),
    ("Platform Metrics (daily)", "Raw daily performance as pulled from Facebook / Google Ads / Reddit / Quora / PulsePoint: impressions, clicks, spend, conversions."),
    ("Salesforce CPV (daily)", "Raw daily CPV (cost-per-visit/value) from Salesforce, matched to adsets via ecd_code. revenue = conversions x cpv."),
    ("Adobe Visits (daily)", "Raw daily site visit counts from Adobe Analytics, matched to adsets via ecd_code."),
    ("Data coverage", f"{db.START_DATE.isoformat()} through {today_str} — this is mock/demo data, and the live app's dataset grows by one more day automatically as time passes."),
]
readme_df = pd.DataFrame(readme_rows[1:], columns=readme_rows[0])

repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
xlsx_path = os.path.join(repo_root, "Sample_Dashboard_Data.xlsx")
csv_path = os.path.join(repo_root, "Sample_Dashboard_Data.csv")

HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")

sheets = {
    "README": readme_df,
    "Unified Metrics": unified,
    "Campaigns": campaigns,
    "Adsets": adsets,
    "Platform Metrics": platform_metrics,
    "Salesforce CPV": salesforce_cpv,
    "Adobe Visits": adobe_visits,
}

with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
    for name, df in sheets.items():
        df.to_excel(writer, sheet_name=name, index=False)
        ws = writer.sheets[name]
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY_FONT
        ws.freeze_panes = "A2"
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            col_letter = col_cells[0].column_letter
            ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 40)

unified.to_csv(csv_path, index=False)

print(f"Wrote {xlsx_path} ({len(unified)} unified rows, {sum(len(df) for df in sheets.values())} total rows across sheets)")
print(f"Wrote {csv_path} ({len(unified)} rows)")
